#!/usr/bin/env python3
"""Convierte planilla_conteo_manual.csv en un .xlsx con formulas VIVAS.

El CSV se importa en LibreOffice/Excel con las formulas como texto salvo que se
marque "evaluar formulas" al abrirlo; el .xlsx no tiene ese problema.

Uso:
  python3 scripts/eval/make_planilla_xlsx.py --dir ~/verificacion_manual_PR
"""
import argparse, csv, os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HDR = ["n", "archivo", "camara", "GT_cajas_verdes", "PRED_cajas_rojas",
       "TP", "FP", "FN", "GT_omitido", "observaciones", "chequeo"]
WIDTHS = [5, 38, 10, 16, 17, 7, 7, 7, 13, 34, 11]
Z = 1.96  # normal al 95%


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", required=True)
    args = ap.parse_args()
    base = os.path.expanduser(args.dir)

    with open(os.path.join(base, "planilla_conteo_manual.csv")) as fh:
        rows = [r for r in csv.reader(fh) if r and r[0].isdigit()]

    wb = Workbook()
    ws = wb.active
    ws.title = "conteo manual"

    thin = Side(style="thin", color="BBBBBB")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_gt = PatternFill("solid", fgColor="D9EAD3")     # verde suave
    fill_pred = PatternFill("solid", fgColor="F4CCCC")   # rojo suave
    fill_in = PatternFill("solid", fgColor="FFF2CC")     # amarillo = se llena a mano
    fill_hdr = PatternFill("solid", fgColor="404040")

    for c, (name, w) in enumerate(zip(HDR, WIDTHS), 1):
        cell = ws.cell(1, c, name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    first, last = 2, 1 + len(rows)
    for i, r in enumerate(rows):
        y = first + i
        ws.cell(y, 1, int(r[0]))
        ws.cell(y, 2, r[1])
        ws.cell(y, 3, r[2])
        ws.cell(y, 4, int(r[3])).fill = fill_gt
        ws.cell(y, 5, int(r[4])).fill = fill_pred
        for c in (6, 7, 8, 9):
            ws.cell(y, c).fill = fill_in
        # TP+FP+GT_omitido debe dar las cajas rojas; TP+FN debe dar las verdes
        ws.cell(y, 11, f'=IF(COUNT(F{y}:I{y})<4,"",'
                       f'IF(AND(F{y}+G{y}+I{y}=E{y},F{y}+H{y}=D{y}),"ok","REVISAR"))')
        for c in range(1, 12):
            ws.cell(y, c).border = box

    tot = last + 1
    ws.cell(tot, 1, "TOTAL").font = Font(bold=True)
    for c in range(4, 10):
        col = get_column_letter(c)
        cell = ws.cell(tot, c, f"=SUM({col}{first}:{col}{last})")
        cell.font = Font(bold=True)
        cell.border = box
    ws.cell(tot, 11, f'=IF(COUNTIF(K{first}:K{last},"REVISAR")>0,"hay filas a revisar","")')

    # --- bloque de resultados: etiqueta en A:B (fusionada), valor en C, IC en D/E ---
    azul = PatternFill("solid", fgColor="DDEBF7")

    def label(row, text):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws.cell(row, 1, text)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="right", vertical="center")

    def value(row, formula):
        c = ws.cell(row, 3, formula)
        c.number_format = "0.000"
        c.font = Font(bold=True, size=12)
        c.fill = azul
        c.border = box
        c.alignment = Alignment(horizontal="center")
        return c

    def metric(row, text, formula, n_formula):
        label(row, text)
        value(row, formula)
        # intervalo de Wilson al 95%
        p, n = f"C{row}", n_formula
        centro = f"(({p}+{Z}^2/(2*{n}))/(1+{Z}^2/{n}))"
        semi = f"({Z}*SQRT({p}*(1-{p})/{n}+{Z}^2/(4*{n}^2))/(1+{Z}^2/{n}))"
        lo = ws.cell(row, 4, f'=IF({n}=0,"",MAX(0,{centro}-{semi}))')
        hi = ws.cell(row, 5, f'=IF({n}=0,"",MIN(1,{centro}+{semi}))')
        for cc in (lo, hi):
            cc.number_format = "0.000"
            cc.alignment = Alignment(horizontal="center")
            cc.border = box

    r_hdr, r_p, r_r, r_f = tot + 2, tot + 3, tot + 4, tot + 5
    ws.cell(r_hdr, 1, "RESULTADO").font = Font(bold=True, size=12)
    for col, txt in ((3, "valor"), (4, "IC 95% inf"), (5, "IC 95% sup")):
        c = ws.cell(r_hdr, col, txt)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
    metric(r_p, "PRECISION =TP/(TP+FP)",
           f'=IF(F{tot}+G{tot}=0,"",F{tot}/(F{tot}+G{tot}))', f"(F{tot}+G{tot})")
    metric(r_r, "RECALL =TP/(TP+FN)",
           f'=IF(F{tot}+H{tot}=0,"",F{tot}/(F{tot}+H{tot}))', f"(F{tot}+H{tot})")
    label(r_f, "F1")
    value(r_f, f'=IF(OR(C{r_p}="",C{r_r}=""),"",2*C{r_p}*C{r_r}/(C{r_p}+C{r_r}))')

    notes = [
        "",
        "Amarillo = se llena a mano (TP / FP / FN / GT_omitido). Lo demas se calcula solo.",
        "TP  = caja roja sobre una cabeza real que tiene su caja verde.",
        "FP  = caja roja que no esta sobre una cabeza real.",
        "FN  = caja verde sin ninguna roja encima (cabeza que el modelo perdio).",
        "GT_omitido = cabeza real detectada por el modelo que el anotador NO marco (no cuenta como FP).",
        "chequeo: TP+FP+GT_omitido debe dar las cajas rojas, y TP+FN las cajas verdes.",
    ]
    for i, t in enumerate(notes):
        ws.cell(r_f + 2 + i, 1, t).font = Font(italic=True, color="555555")

    out = os.path.join(base, "planilla_conteo_manual.xlsx")
    wb.save(out)
    print("OK ->", out)


if __name__ == "__main__":
    main()
